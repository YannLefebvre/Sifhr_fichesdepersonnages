#!/usr/bin/env python3
"""
update_traits.py — Met à jour TRAITS_DATA dans octogone.html depuis le tableur Excel.

Usage :
    python3 update_traits.py Traits_Prodiges_Artefacts.xlsx octogone.html

Le script remplace uniquement la ligne "const TRAITS_DATA = [...];"
sans toucher au reste du fichier.
"""

import sys
import json
import re

try:
    from openpyxl import load_workbook
except ImportError:
    print("Erreur : openpyxl non installé. Lancez : pip install openpyxl")
    sys.exit(1)

if len(sys.argv) < 3:
    print("Usage : python3 update_traits.py <fichier.xlsx> <octogone.html>")
    sys.exit(1)

xlsx_path = sys.argv[1]
html_path = sys.argv[2]

# ── Lire le tableur ──
def parse_niveaux(val):
    if not val: return [1,2,3]
    v = str(val).strip()
    if v == 'Variable': return [1,2,3]
    if v in ('N1','1'): return [1]
    if v in ('N2','2'): return [2]
    if v in ('N3','3'): return [3]
    if v in ('N1 à N3','1 à 3'): return [1,2,3]
    if v in ('N1 et 2','1 et 2','N1 ou N2'): return [1,2]
    if v in ('N2 à N3','2 à 3','N2 ou N3'): return [2,3]
    return [1,2,3]

wb = load_workbook(xlsx_path)
ws = wb.active
traits = []
for r in range(2, ws.max_row+1):
    nom = ws.cell(r,1).value
    if not nom: continue
    traits.append({
        'nom':       str(nom).strip(),
        'niveaux':   parse_niveaux(ws.cell(r,2).value),
        'type':      str(ws.cell(r,3).value or '').strip(),
        'descriptif':str(ws.cell(r,4).value or '').strip(),
        'aptitudes': [a.strip() for a in str(ws.cell(r,5).value or '').split(',') if a.strip()],
        'tags':      [t.strip() for t in str(ws.cell(r,6).value or '').split(',') if t.strip()],
        'categorie': str(ws.cell(r,7).value or 'Trait').strip()
    })

print(f"✓ {len(traits)} traits lus depuis {xlsx_path}")

# ── Remplacer dans le HTML ──
new_line = "const TRAITS_DATA = " + json.dumps(traits, ensure_ascii=False, separators=(',',':')) + ";"

with open(html_path,'r',encoding='utf-8') as f:
    html = f.read()

pattern = r'const TRAITS_DATA\s*=\s*\[.*?\];'
if not re.search(pattern, html, re.DOTALL):
    print("Erreur : 'const TRAITS_DATA' introuvable dans le fichier HTML.")
    sys.exit(1)

html_updated = re.sub(pattern, new_line, html, flags=re.DOTALL)

with open(html_path,'w',encoding='utf-8') as f:
    f.write(html_updated)

print(f"✓ TRAITS_DATA mis à jour dans {html_path}")
print(f"  (ancienne taille: {len(html)} chars → nouvelle: {len(html_updated)} chars)")
